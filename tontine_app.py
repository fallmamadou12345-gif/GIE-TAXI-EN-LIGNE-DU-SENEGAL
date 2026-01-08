from __future__ import annotations

import os
import json
import sqlite3
from datetime import date, datetime
from typing import Optional, List, Tuple, Dict

from fastapi import FastAPI, Request, Form, Response
from fastapi.responses import HTMLResponse, RedirectResponse, PlainTextResponse
from jinja2 import Environment, BaseLoader
from passlib.hash import pbkdf2_sha256

# ============================================================
# CONFIG
# ============================================================
APP_TITLE = "Tontine"
DB_PATH = os.environ.get("TONTINE_DB", "tontine.db")
ADMIN_PASSWORD = os.environ.get("TONTINE_ADMIN_PASSWORD", "admin1234")

DEFAULT_START_DATE = "2025-10-15"
DEFAULT_DAILY_AMOUNT = "2000"          # 2000 F / jour
DEFAULT_MONTHLY_CAP = "60000"          # 60 000 F / mois (plafond)
DEFAULT_OCT_2025_FIXED = "30000"       # Oct 2025 = 30 000 F (forfait)
DEFAULT_END_DATE = ""                  # vide = pas d'arrêt


# ============================================================
# DB
# ============================================================
def db() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def init_db() -> None:
    with db() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                phone TEXT NOT NULL UNIQUE,
                pin_hash TEXT NOT NULL,
                start_date TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                member_id INTEGER NOT NULL,
                pay_date TEXT NOT NULL,
                amount INTEGER NOT NULL,
                mode TEXT NOT NULL,
                reference TEXT,
                note TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(member_id) REFERENCES members(id)
            );

            CREATE INDEX IF NOT EXISTS idx_payments_member_date
            ON payments(member_id, pay_date);

            CREATE TABLE IF NOT EXISTS bank_deposits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dep_date TEXT NOT NULL,
                amount INTEGER NOT NULL,
                reference TEXT,
                note TEXT,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_bank_deposits_date
            ON bank_deposits(dep_date);
            """
        )

        defaults = {
            "project_start_date": DEFAULT_START_DATE,
            "daily_amount": DEFAULT_DAILY_AMOUNT,
            "monthly_cap": DEFAULT_MONTHLY_CAP,
            "oct_2025_fixed": DEFAULT_OCT_2025_FIXED,
            "project_end_date": DEFAULT_END_DATE,
        }
        for k, v in defaults.items():
            row = con.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone()
            if not row:
                con.execute("INSERT INTO settings(key,value) VALUES(?,?)", (k, v))


# ============================================================
# Settings helpers
# ============================================================
def get_setting(key: str, default: str = "") -> str:
    with db() as con:
        row = con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with db() as con:
        con.execute(
            "INSERT INTO settings(key,value) VALUES(?,?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )


def project_start_date() -> date:
    return date.fromisoformat(get_setting("project_start_date", DEFAULT_START_DATE))


def daily_amount() -> int:
    return int(get_setting("daily_amount", DEFAULT_DAILY_AMOUNT))


def monthly_cap() -> int:
    return int(get_setting("monthly_cap", DEFAULT_MONTHLY_CAP))


def oct_2025_fixed() -> int:
    return int(get_setting("oct_2025_fixed", DEFAULT_OCT_2025_FIXED))


def project_end_date() -> Optional[date]:
    s = get_setting("project_end_date", DEFAULT_END_DATE).strip()
    return date.fromisoformat(s) if s else None


# ============================================================
# Helpers
# ============================================================
def today() -> date:
    return date.today()


def as_of_date() -> date:
    """Date de calcul = min(today, date d'arrêt) si le projet est arrêté."""
    endd = project_end_date()
    t = today()
    return min(t, endd) if endd else t


def is_valid_sn_phone(phone: str) -> bool:
    return phone.isdigit() and len(phone) == 9


def month_end(d: date) -> date:
    if d.month == 12:
        return date(d.year, 12, 31)
    nxt = date(d.year, d.month + 1, 1)
    return nxt.fromordinal(nxt.toordinal() - 1)


def months_between(start: date, end: date) -> List[Tuple[int, int]]:
    y, m = start.year, start.month
    out: List[Tuple[int, int]] = []
    while (y, m) <= (end.year, end.month):
        out.append((y, m))
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1
    return out


def due_for_month(year: int, month: int, asof: date, member_start: date) -> int:
    # Avant l'adhésion
    if (year, month) < (member_start.year, member_start.month):
        return 0

    # Octobre 2025 forfait (selon ton besoin)
    if year == 2025 and month == 10:
        if (member_start.year, member_start.month) != (2025, 10):
            return 0
        # Forfait compté si on a atteint la fin d'octobre (ou après)
        if asof < date(2025, 10, 31):
            return 0
        return oct_2025_fixed()

    m_start = date(year, month, 1)
    m_end = month_end(m_start)

    eff_start = max(m_start, member_start)
    eff_end = min(m_end, asof)
    if eff_end < eff_start:
        return 0

    days = (eff_end - eff_start).days + 1
    return min(monthly_cap(), daily_amount() * days)


def due_total_as_of(asof: date, member_start: date) -> int:
    if asof < member_start:
        return 0
    total = 0
    for y, m in months_between(member_start, asof):
        total += due_for_month(y, m, asof, member_start)
    return total


def paid_total_as_of(member_id: int, asof: date) -> int:
    with db() as con:
        row = con.execute(
            "SELECT COALESCE(SUM(amount),0) AS s FROM payments WHERE member_id=? AND pay_date<=?",
            (member_id, asof.isoformat()),
        ).fetchone()
        return int(row["s"])


def member_status(member_id: int, member_start: date, asof: date) -> Dict[str, int | str]:
    due = due_total_as_of(asof, member_start)
    paid = paid_total_as_of(member_id, asof)
    rest = due - paid
    balance = paid - due
    avance = max(0, balance)
    rest_pos = max(0, -balance)
    status = "À jour" if rest_pos == 0 else "En retard"
    return {
        "due": due,
        "paid": paid,
        "rest": rest_pos,
        "avance": avance,
        "balance": balance,
        "status": status,
    }


def total_collected_as_of(asof: date) -> int:
    with db() as con:
        row = con.execute(
            "SELECT COALESCE(SUM(amount),0) AS s FROM payments WHERE pay_date<=?",
            (asof.isoformat(),),
        ).fetchone()
        return int(row["s"])


def total_deposited_as_of(asof: date) -> int:
    with db() as con:
        row = con.execute(
            "SELECT COALESCE(SUM(amount),0) AS s FROM bank_deposits WHERE dep_date<=?",
            (asof.isoformat(),),
        ).fetchone()
        return int(row["s"])


def fmt(n: int) -> str:
    return f"{int(n):,}".replace(",", " ")


def ym_list(start: date, end: date) -> List[str]:
    out = []
    for y, m in months_between(start, end):
        out.append(f"{y:04d}-{m:02d}")
    return out


def monthly_collection_stats(asof: date) -> Dict[str, Dict[str, int]]:
    """
    {
      "YYYY-MM": {"total":..., "wave":..., "om":..., "cash":...}
    }
    """
    start = project_start_date()
    end = asof
    months = ym_list(start, end)
    stats = {k: {"total": 0, "wave": 0, "om": 0, "cash": 0} for k in months}

    with db() as con:
        rows = con.execute(
            """
            SELECT substr(pay_date,1,7) AS ym,
                   COALESCE(SUM(amount),0) AS total,
                   COALESCE(SUM(CASE WHEN mode='wave' THEN amount ELSE 0 END),0) AS wave_total,
                   COALESCE(SUM(CASE WHEN mode='om' THEN amount ELSE 0 END),0) AS om_total,
                   COALESCE(SUM(CASE WHEN mode='cash' THEN amount ELSE 0 END),0) AS cash_total
            FROM payments
            WHERE pay_date <= ?
            GROUP BY substr(pay_date,1,7)
            ORDER BY ym
            """,
            (asof.isoformat(),),
        ).fetchall()

    for r in rows:
        ym = r["ym"]
        if ym not in stats:
            stats[ym] = {"total": 0, "wave": 0, "om": 0, "cash": 0}
        stats[ym]["total"] = int(r["total"])
        stats[ym]["wave"] = int(r["wave_total"])
        stats[ym]["om"] = int(r["om_total"])
        stats[ym]["cash"] = int(r["cash_total"])

    return stats


# ============================================================
# HTML / THEME PREMIUM + MODE CLAIR/SOMBRE + TABLES + SEARCH
# ============================================================
env = Environment(loader=BaseLoader())

BASE = """
<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{{ title }}</title>
  <style>
    :root{
      --bg:#0b1220;
      --panel:#0f1a30;
      --card:#111f3a;
      --text:#e8eefc;
      --muted:#9fb1d6;
      --line:rgba(255,255,255,.08);
      --brand:#6aa8ff;
      --brand2:#9b7bff;
      --good:#2dd4bf;
      --bad:#fb7185;
      --warn:#fbbf24;
      --shadow: 0 12px 40px rgba(0,0,0,.35);
      --radius:18px;
    }

    body[data-theme="light"]{
      --bg:#f5f7ff;
      --panel:#ffffff;
      --card:#ffffff;
      --text:#0a1020;
      --muted:#4b5b7e;
      --line:rgba(10,16,32,.10);
      --shadow: 0 10px 28px rgba(10,16,32,.12);
      background:
        radial-gradient(1200px 600px at 10% 0%, rgba(106,168,255,.22), transparent 55%),
        radial-gradient(1000px 700px at 95% 15%, rgba(155,123,255,.16), transparent 55%),
        linear-gradient(180deg, #ffffff, var(--bg));
    }

    *{box-sizing:border-box}
    body{
      margin:0;
      font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial, "Noto Sans";
      color:var(--text);
      background:
        radial-gradient(1200px 600px at 10% 0%, rgba(106,168,255,.25), transparent 55%),
        radial-gradient(1000px 700px at 95% 15%, rgba(155,123,255,.20), transparent 55%),
        radial-gradient(900px 600px at 50% 100%, rgba(45,212,191,.10), transparent 55%),
        linear-gradient(180deg, #070b15, var(--bg));
      min-height:100vh;
    }

    a{color:var(--text)}
    body[data-theme="light"] a{color:var(--text)}
    .muted{color:var(--muted); font-size:13px}
    .wrap{max-width:1200px;margin:0 auto;padding:18px}

    .topbar{
      display:flex; align-items:center; justify-content:space-between; gap:12px;
      background:rgba(17,31,58,.55);
      border:1px solid var(--line);
      backdrop-filter: blur(10px);
      padding:12px 14px;
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      position:sticky; top:10px; z-index:20;
    }
    body[data-theme="light"] .topbar{background:rgba(255,255,255,.75)}
    .brand{
      display:flex; align-items:center; gap:10px;
    }
    .logo{
      width:36px;height:36px;border-radius:12px;
      background: linear-gradient(135deg, var(--brand), var(--brand2));
      box-shadow: 0 10px 30px rgba(106,168,255,.25);
    }
    .brand strong{font-size:15px; letter-spacing:.2px}
    .menu{display:flex; flex-wrap:wrap; gap:8px; align-items:center}

    .pill{
      display:inline-flex; align-items:center; gap:8px;
      padding:8px 10px;
      border-radius:999px;
      border:1px solid var(--line);
      background:rgba(15,26,48,.6);
      text-decoration:none;
      font-size:13px;
      transition: transform .08s ease, background .15s ease, border-color .15s ease;
      cursor:pointer;
      color:var(--text);
    }
    body[data-theme="light"] .pill{background:rgba(255,255,255,.75)}
    .pill:hover{transform: translateY(-1px); background:rgba(15,26,48,.85); border-color:rgba(255,255,255,.14)}
    body[data-theme="light"] .pill:hover{background:rgba(255,255,255,.95)}
    .pill .dot{width:8px;height:8px;border-radius:999px;background:rgba(255,255,255,.35)}
    body[data-theme="light"] .pill .dot{background:rgba(10,16,32,.35)}
    .pill.primary{border-color: rgba(106,168,255,.30)}
    .pill.primary .dot{background: var(--brand)}
    .pill.danger{border-color: rgba(251,113,133,.35)}
    .pill.danger .dot{background: var(--bad)}

    .card{
      background: rgba(17,31,58,.65);
      border:1px solid var(--line);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding:16px;
      margin-bottom:14px;
    }
    body[data-theme="light"] .card{background: rgba(255,255,255,.90)}
    .card h1{font-size:20px;margin:0 0 10px}
    .card h2{font-size:16px;margin:0 0 10px}

    .row{display:flex; gap:12px; flex-wrap:wrap}
    .col{flex:1; min-width:240px}

    label{display:block; font-size:12px; color:var(--muted); margin-bottom:6px}
    input, select{
      width:100%;
      padding:11px 12px;
      border-radius: 14px;
      border:1px solid rgba(255,255,255,.10);
      background: rgba(8,12,24,.55);
      color:var(--text);
      outline:none;
      transition:border-color .15s ease, transform .08s ease;
    }
    body[data-theme="light"] input, body[data-theme="light"] select{
      background: rgba(245,247,255,.85);
      border:1px solid rgba(10,16,32,.12);
      color: var(--text);
    }
    input:focus, select:focus{
      border-color: rgba(106,168,255,.45);
      box-shadow: 0 0 0 4px rgba(106,168,255,.12);
    }
    input::placeholder{color: rgba(159,177,214,.55)}
    body[data-theme="light"] input::placeholder{color: rgba(75,91,126,.55)}

    .btn{
      display:inline-flex; align-items:center; justify-content:center; gap:8px;
      padding:11px 14px;
      border-radius: 14px;
      border:1px solid rgba(255,255,255,.10);
      cursor:pointer;
      text-decoration:none;
      font-weight:700;
      transition: transform .08s ease, background .15s ease, border-color .15s ease;
      user-select:none;
      color:var(--text);
      background: rgba(15,26,48,.55);
    }
    body[data-theme="light"] .btn{background: rgba(245,247,255,.90); border:1px solid rgba(10,16,32,.12)}
    .btn:hover{transform: translateY(-1px)}
    .btn:active{transform: translateY(0px)}
    .btn.primary{
      background: linear-gradient(135deg, rgba(106,168,255,.95), rgba(155,123,255,.95));
      border-color: rgba(106,168,255,.35);
      color:#061022;
    }
    .btn.danger{
      background: rgba(251,113,133,.16);
      border-color: rgba(251,113,133,.30);
    }

    .kpis{
      display:grid;
      grid-template-columns: repeat(4, 1fr);
      gap:12px;
    }
    @media(max-width:1100px){ .kpis{grid-template-columns: repeat(2, 1fr)} }
    @media(max-width:520px){ .kpis{grid-template-columns: 1fr} }

    .kpi{
      background: rgba(14,24,48,.55);
      border:1px solid var(--line);
      border-radius: 16px;
      padding:12px;
      position:relative;
      overflow:hidden;
    }
    body[data-theme="light"] .kpi{background: rgba(245,247,255,.80)}
    .kpi:before{
      content:"";
      position:absolute; inset:-2px;
      background: radial-gradient(400px 120px at 15% 0%, rgba(106,168,255,.18), transparent 60%);
      pointer-events:none;
    }
    .kpi .t{font-size:12px;color:var(--muted);margin-bottom:6px}
    .kpi .v{font-size:20px;font-weight:900;letter-spacing:.2px}

    .badge{
      display:inline-flex; align-items:center; gap:8px;
      padding:7px 10px;
      border-radius:999px;
      font-size:12px;
      border:1px solid var(--line);
      background: rgba(8,12,24,.35);
    }
    body[data-theme="light"] .badge{background: rgba(245,247,255,.85)}
    .badge .b{width:8px;height:8px;border-radius:999px}
    .badge.ok .b{background: var(--good)}
    .badge.bad .b{background: var(--bad)}
    .badge.warn .b{background: var(--warn)}

    .table-wrap{
      border:1px solid var(--line);
      border-radius: 16px;
      overflow:hidden;
      background: rgba(8,12,24,.25);
    }
    body[data-theme="light"] .table-wrap{background: rgba(245,247,255,.75)}
    table{width:100%; border-collapse:collapse}
    th, td{
      padding:12px 12px;
      border-bottom:1px solid rgba(255,255,255,.06);
      text-align:left;
      font-size:13px;
      vertical-align:top;
    }
    body[data-theme="light"] th, body[data-theme="light"] td{border-bottom:1px solid rgba(10,16,32,.08)}
    th{
      font-size:12px;
      color:rgba(159,177,214,.95);
      background: rgba(15,26,48,.65);
      cursor:pointer;
      user-select:none;
      position:sticky;
      top:0;
      z-index:5;
    }
    body[data-theme="light"] th{color: rgba(75,91,126,.95); background: rgba(245,247,255,.95)}
    tr:hover td{background: rgba(106,168,255,.06)}
    .mono{font-variant-numeric: tabular-nums}

    .searchbar{
      display:flex; gap:10px; flex-wrap:wrap; align-items:flex-end;
      margin:10px 0 12px;
    }

    .warnbox{
      background: rgba(251,191,36,.10);
      border: 1px solid rgba(251,191,36,.25);
      border-radius: 16px;
      padding: 12px;
    }

    .toast{
      position: fixed;
      right: 14px;
      bottom: 14px;
      background: rgba(15,26,48,.90);
      border:1px solid var(--line);
      border-radius: 14px;
      padding: 12px 14px;
      box-shadow: var(--shadow);
      display:none;
      max-width: 360px;
    }
    body[data-theme="light"] .toast{background: rgba(255,255,255,.95)}
  </style>
</head>

<body>
  <div class="wrap">
    <div class="topbar">
      <div class="brand">
        <div class="logo"></div>
        <div>
          <div><strong>{{ app_title }}</strong></div>
          <div class="muted">
            Début: {{ pstart }}
            {% if pend %} · Arrêt: {{ pend }}{% endif %}
            · Calcul: {{ asof }}
          </div>
        </div>
      </div>

      <div class="menu">
        {% if admin %}
          <a class="pill primary" href="/admin"><span class="dot"></span>Dashboard</a>
          <a class="pill" href="/admin/resume"><span class="dot"></span>Résumé</a>
          <a class="pill" href="/admin/comptabilite"><span class="dot"></span>Comptabilité</a>
          <a class="pill" href="/admin/restes"><span class="dot"></span>Retards</a>
          <a class="pill" href="/admin/config"><span class="dot"></span>Configuration</a>
        {% endif %}
        {% if member %}
          <a class="pill primary" href="/me"><span class="dot"></span>Mon profil</a>
        {% endif %}

        <button class="pill" type="button" onclick="toggleTheme()" title="Mode clair/sombre">
          <span class="dot"></span>Thème
        </button>

        {% if admin or member %}
          <a class="pill danger" href="/logout"><span class="dot"></span>Déconnexion</a>
        {% endif %}
      </div>
    </div>

    <div style="height:14px"></div>

    {{ body|safe }}

    <div class="muted" style="margin-top:14px;">
      Modes autorisés: Wave / Orange Money / Cash. Aucune pénalité.
    </div>
  </div>

  <div id="toast" class="toast"></div>

  <script>
    function toast(msg){
      const t = document.getElementById("toast");
      if(!t) return;
      t.textContent = msg;
      t.style.display = "block";
      clearTimeout(window.__toastTimer);
      window.__toastTimer = setTimeout(()=>{ t.style.display="none"; }, 2600);
    }

    // Search in tables: input[data-search-table="#id"]
    document.querySelectorAll("input[data-search-table]").forEach(inp=>{
      inp.addEventListener("input", ()=>{
        const sel = inp.getAttribute("data-search-table");
        const q = (inp.value||"").toLowerCase().trim();
        const table = document.querySelector(sel);
        if(!table) return;
        table.querySelectorAll("tbody tr").forEach(tr=>{
          const text = tr.innerText.toLowerCase();
          tr.style.display = text.includes(q) ? "" : "none";
        });
      });
    });

    // Sort table by clicking headers
    document.querySelectorAll("table[data-sortable]").forEach(table=>{
      table.querySelectorAll("th").forEach((th, idx)=>{
        th.addEventListener("click", ()=>{
          const tbody = table.querySelector("tbody");
          const rows = Array.from(tbody.querySelectorAll("tr"));
          const asc = th.getAttribute("data-asc") !== "1";
          table.querySelectorAll("th").forEach(x=>x.removeAttribute("data-asc"));
          th.setAttribute("data-asc", asc ? "1" : "0");

          rows.sort((a,b)=>{
            const ta = (a.children[idx]?.innerText||"").trim();
            const tb = (b.children[idx]?.innerText||"").trim();
            const na = parseFloat(ta.replace(/[^0-9.-]/g,""));
            const nb = parseFloat(tb.replace(/[^0-9.-]/g,""));
            const bothNum = !isNaN(na) && !isNaN(nb);
            if(bothNum) return asc ? (na-nb) : (nb-na);
            return asc ? ta.localeCompare(tb) : tb.localeCompare(ta);
          });

          rows.forEach(r=>tbody.appendChild(r));
          toast("Tri appliqué");
        });
      });
    });

    function applyTheme(t){
      document.body.setAttribute("data-theme", t);
      localStorage.setItem("tontine_theme", t);
      toast("Thème: " + (t === "light" ? "Clair" : "Sombre"));
    }
    function toggleTheme(){
      const cur = document.body.getAttribute("data-theme") || "dark";
      applyTheme(cur === "dark" ? "light" : "dark");
    }
    (function(){
      const saved = localStorage.getItem("tontine_theme") || "dark";
      document.body.setAttribute("data-theme", saved);
    })();
  </script>
</body>
</html>
"""


def render(title: str, body: str, admin: bool = False, member: bool = False) -> HTMLResponse:
    tpl = env.from_string(BASE)
    html = tpl.render(
        title=title,
        body=body,
        admin=admin,
        member=member,
        app_title=APP_TITLE,
        asof=as_of_date().isoformat(),
        pstart=project_start_date().isoformat(),
        pend=project_end_date().isoformat() if project_end_date() else "",
    )
    return HTMLResponse(html)


# ============================================================
# Cookies
# ============================================================
def set_cookie(resp: Response, key: str, value: str) -> None:
    resp.set_cookie(key=key, value=value, httponly=True, samesite="lax")


def clear_cookie(resp: Response, key: str) -> None:
    resp.delete_cookie(key)


def get_cookie(request: Request, key: str) -> Optional[str]:
    return request.cookies.get(key)


# ============================================================
# App
# ============================================================
app = FastAPI(title=APP_TITLE)


@app.on_event("startup")
def _startup():
    init_db()


@app.get("/favicon.ico")
def favicon():
    return Response(status_code=204)


# ============================================================
# Auth guards
# ============================================================
def require_admin(request: Request) -> Optional[HTMLResponse]:
    if get_cookie(request, "admin") != "1":
        return render("Accès refusé", '<div class="card"><h1>Accès refusé</h1><a class="btn" href="/">Retour</a></div>')
    return None


def require_member(request: Request) -> tuple[Optional[HTMLResponse], Optional[int]]:
    mid = get_cookie(request, "member_id")
    if not mid or not mid.isdigit():
        return render("Accès refusé", '<div class="card"><h1>Accès refusé</h1><a class="btn" href="/">Retour</a></div>'), None
    return None, int(mid)


# ============================================================
# Home + login
# ============================================================
@app.get("/", response_class=HTMLResponse)
def home():
    body = """
    <div class="card">
      <h1>Bienvenue</h1>
      <div class="muted">Connecte-toi en tant que Membre ou Admin.</div>
      <div style="height:12px"></div>

      <div class="row">
        <div class="col card">
          <h2>Accès Membre</h2>
          <form method="post" action="/login-member">
            <label>Téléphone (9 chiffres)</label>
            <input name="phone" placeholder="77xxxxxxx" required />
            <div style="height:10px"></div>
            <label>PIN (4 chiffres)</label>
            <input name="pin" placeholder="1234" required />
            <div style="height:12px"></div>
            <button class="btn primary" type="submit">Se connecter</button>
          </form>
        </div>

        <div class="col card">
          <h2>Accès Admin</h2>
          <form method="post" action="/login-admin">
            <label>Mot de passe admin</label>
            <input name="password" type="password" required />
            <div style="height:12px"></div>
            <button class="btn primary" type="submit">Connexion Admin</button>
          </form>
          <div class="muted" style="margin-top:10px;">Changer via variable: <b>TONTINE_ADMIN_PASSWORD</b></div>
        </div>
      </div>
    </div>
    """
    return render("Accueil", body)


@app.get("/logout")
def logout():
    resp = RedirectResponse("/", status_code=303)
    clear_cookie(resp, "admin")
    clear_cookie(resp, "member_id")
    return resp


@app.post("/login-admin")
def login_admin(password: str = Form(...)):
    if password != ADMIN_PASSWORD:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Mot de passe admin incorrect.</p><a class="btn" href="/">Retour</a></div>')
    resp = RedirectResponse("/admin", status_code=303)
    set_cookie(resp, "admin", "1")
    return resp


@app.post("/login-member")
def login_member(phone: str = Form(...), pin: str = Form(...)):
    phone = phone.strip()
    pin = pin.strip()

    if not is_valid_sn_phone(phone):
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Téléphone invalide (9 chiffres).</p><a class="btn" href="/">Retour</a></div>')

    if not (pin.isdigit() and len(pin) == 4):
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>PIN invalide (4 chiffres).</p><a class="btn" href="/">Retour</a></div>')

    with db() as con:
        m = con.execute("SELECT * FROM members WHERE phone=?", (phone,)).fetchone()

    if not m or int(m["is_active"]) != 1:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Compte désactivé ou introuvable.</p><a class="btn" href="/">Retour</a></div>')

    if not pbkdf2_sha256.verify(pin, m["pin_hash"]):
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Téléphone ou PIN incorrect.</p><a class="btn" href="/">Retour</a></div>')

    resp = RedirectResponse("/me", status_code=303)
    set_cookie(resp, "member_id", str(m["id"]))
    return resp


# ============================================================
# Admin Dashboard
# ============================================================
@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request):
    gate = require_admin(request)
    if gate:
        return gate

    asof = as_of_date()
    pstart = project_start_date()

    with db() as con:
        members = con.execute("SELECT * FROM members ORDER BY full_name").fetchall()

    rows = []
    total_due = total_paid = total_rest = total_avance = 0
    ok_count = late_count = 0
    active_count = 0

    for m in members:
        is_active = int(m["is_active"]) == 1
        if is_active:
            active_count += 1

        m_start = date.fromisoformat(m["start_date"])
        st = member_status(int(m["id"]), m_start, asof) if is_active else {"due": 0, "paid": 0, "rest": 0, "avance": 0, "status": "Désactivé"}

        if is_active:
            total_due += int(st["due"])
            total_paid += int(st["paid"])
            total_rest += int(st["rest"])
            total_avance += int(st["avance"])
            if st["status"] == "À jour":
                ok_count += 1
                badge = "ok"
                label = "À jour"
            else:
                late_count += 1
                badge = "bad"
                label = "En retard"
        else:
            badge = "bad"
            label = "Désactivé"

        rows.append(f"""
          <tr>
            <td>
              <a href="/admin/member/{m['id']}"><b>{m['full_name']}</b></a>
              <div class="muted">{m['phone']}</div>
            </td>
            <td><span class="badge {badge}"><span class="b"></span>{label}</span></td>
            <td class="mono">{fmt(int(st['paid']))} F</td>
            <td class="mono">{fmt(int(st['due']))} F</td>
            <td class="mono"><b>{fmt(int(st['rest']))} F</b></td>
            <td class="mono">{fmt(int(st['avance']))} F</td>
          </tr>
        """)

    collected = total_collected_as_of(asof)
    deposited = total_deposited_as_of(asof)
    available_bank = max(0, collected - deposited)

    pend = project_end_date()
    stop_banner = ""
    if pend:
        stop_banner = f"""
        <div class="card">
          <div class="warnbox">
            <b>Projet arrêté</b> : les calculs sont bloqués à <b>{pend.isoformat()}</b>.
            <div style="height:10px"></div>
            <form method="post" action="/admin/project/resume" onsubmit="return confirm('Reprendre le projet ?');">
              <button class="btn" type="submit">Reprendre le projet</button>
            </form>
          </div>
        </div>
        """

    body = f"""
    {stop_banner}

    <div class="card">
      <h1>Dashboard Admin</h1>
      <div class="muted">Début projet: <b>{pstart.isoformat()}</b> · Journalier: <b>{fmt(daily_amount())} F</b> · Plafond/mois: <b>{fmt(monthly_cap())} F</b> · Oct 2025: <b>{fmt(oct_2025_fixed())} F</b></div>
      <div style="height:14px"></div>

      <div class="kpis">
        <div class="kpi"><div class="t">Participants</div><div class="v">{len(members)} (actifs {active_count})</div></div>
        <div class="kpi"><div class="t">À jour / En retard</div><div class="v">{ok_count} / {late_count}</div></div>
        <div class="kpi"><div class="t">Total Payé (actifs)</div><div class="v">{fmt(total_paid)} F</div></div>
        <div class="kpi"><div class="t">Total Reste (actifs)</div><div class="v">{fmt(total_rest)} F</div></div>
      </div>

      <div style="height:14px"></div>
      <div class="kpis">
        <div class="kpi"><div class="t">Total encaissé (début → {asof.isoformat()})</div><div class="v">{fmt(collected)} F</div></div>
        <div class="kpi"><div class="t">Total versé banque</div><div class="v">{fmt(deposited)} F</div></div>
        <div class="kpi"><div class="t">Solde dispo à verser</div><div class="v">{fmt(available_bank)} F</div></div>
        <div class="kpi"><div class="t">Résumé projet</div><div class="v"><a href="/admin/resume">Ouvrir</a></div></div>
      </div>
    </div>

    <div class="card">
      <h2>Ajouter un membre</h2>
      <form method="post" action="/admin/add-member">
        <div class="row">
          <div class="col"><label>Nom complet</label><input name="full_name" required /></div>
          <div class="col"><label>Téléphone (9 chiffres)</label><input name="phone" placeholder="77xxxxxxx" required /></div>
          <div class="col"><label>PIN (4 chiffres)</label><input name="pin" placeholder="1234" required /></div>
        </div>
        <div style="height:10px"></div>
        <label>Date de début du membre (AAAA-MM-JJ)</label>
        <input name="start_date" value="{project_start_date().isoformat()}" />
        <div style="height:12px"></div>
        <button class="btn primary" type="submit">Créer membre</button>
        <a class="btn" href="/admin/config">Configuration</a>
      </form>
    </div>

    <div class="card">
      <h2>Liste des membres</h2>

      <div class="searchbar">
        <div style="flex:1; min-width:240px;">
          <label>Recherche (nom / téléphone)</label>
          <input placeholder="Tape ici..." data-search-table="#tblMembers"/>
        </div>
      </div>

      <div class="table-wrap">
        <table id="tblMembers" data-sortable>
          <thead><tr><th>Membre</th><th>Statut</th><th>Payé</th><th>Dû</th><th>Reste</th><th>Avance</th></tr></thead>
          <tbody>{''.join(rows) if rows else '<tr><td colspan="6">Aucun membre</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    """
    return render("Admin", body, admin=True)


@app.post("/admin/add-member")
def admin_add_member(
    request: Request,
    full_name: str = Form(...),
    phone: str = Form(...),
    pin: str = Form(...),
    start_date: str = Form(...)
):
    gate = require_admin(request)
    if gate:
        return gate

    full_name = full_name.strip()
    phone = phone.strip()
    pin = pin.strip()
    start_date = start_date.strip()

    if not is_valid_sn_phone(phone):
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Téléphone invalide (9 chiffres).</p><a class="btn" href="/admin">Retour</a></div>', admin=True)
    if not (pin.isdigit() and len(pin) == 4):
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>PIN invalide (4 chiffres).</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    try:
        sd = date.fromisoformat(start_date)
    except Exception:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Date invalide (AAAA-MM-JJ).</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    with db() as con:
        try:
            con.execute(
                "INSERT INTO members(full_name, phone, pin_hash, start_date, is_active, created_at) VALUES(?,?,?,?,1,?)",
                (full_name, phone, pbkdf2_sha256.hash(pin), sd.isoformat(), datetime.now().isoformat(timespec="seconds")),
            )
        except sqlite3.IntegrityError:
            return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Ce téléphone existe déjà.</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    return RedirectResponse("/admin", status_code=303)


# ============================================================
# Admin member detail + update + change PIN + activate/deactivate + add payment
# ============================================================
@app.get("/admin/member/{member_id}", response_class=HTMLResponse)
def admin_member(request: Request, member_id: int):
    gate = require_admin(request)
    if gate:
        return gate

    asof = as_of_date()

    with db() as con:
        m = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
        if not m:
            return render("Introuvable", '<div class="card"><h1>Membre introuvable</h1><a class="btn" href="/admin">Retour</a></div>', admin=True)

        pays = con.execute(
            "SELECT * FROM payments WHERE member_id=? ORDER BY pay_date DESC, id DESC LIMIT 400",
            (member_id,),
        ).fetchall()

    is_active = int(m["is_active"]) == 1
    m_start = date.fromisoformat(m["start_date"])
    st = member_status(member_id, m_start, asof) if is_active else {"due": 0, "paid": 0, "rest": 0, "avance": 0, "status": "Désactivé"}

    badge_cls = "ok" if is_active and st["status"] == "À jour" else "bad"
    badge_text = ("Actif · " + st["status"]) if is_active else "Désactivé"

    pay_rows = []
    for p in pays:
        pay_rows.append(f"""
          <tr>
            <td class="mono">{p['pay_date']}</td>
            <td class="mono"><b>{fmt(int(p['amount']))} F</b></td>
            <td>{p['mode']}</td>
            <td class="muted">{p['reference'] or ""}</td>
            <td class="muted">{p['note'] or ""}</td>
          </tr>
        """)

    if is_active:
        action_btn = f"""
        <form method="post" action="/admin/member/deactivate" onsubmit="return confirm('Désactiver ce participant ?');">
          <input type="hidden" name="member_id" value="{member_id}" />
          <button class="btn danger" type="submit">Désactiver participant</button>
        </form>
        """
    else:
        action_btn = f"""
        <form method="post" action="/admin/member/activate" onsubmit="return confirm('Réactiver ce participant ?');">
          <input type="hidden" name="member_id" value="{member_id}" />
          <button class="btn" type="submit">Réactiver participant</button>
        </form>
        """

    body = f"""
    <div class="card">
      <h1>{m['full_name']}</h1>
      <div class="muted">Téléphone: <b>{m['phone']}</b> · Début membre: <b>{m['start_date']}</b></div>
      <div style="height:10px"></div>
      <span class="badge {badge_cls}"><span class="b"></span>{badge_text}</span>

      <div style="height:14px"></div>
      <div class="kpis">
        <div class="kpi"><div class="t">Payé</div><div class="v">{fmt(int(st['paid']))} F</div></div>
        <div class="kpi"><div class="t">Dû (début → {asof.isoformat()})</div><div class="v">{fmt(int(st['due']))} F</div></div>
        <div class="kpi"><div class="t">Reste</div><div class="v">{fmt(int(st['rest']))} F</div></div>
        <div class="kpi"><div class="t">Avance</div><div class="v">{fmt(int(st['avance']))} F</div></div>
      </div>
    </div>

    <div class="card">
      <h2>Actions Admin</h2>
      <div class="row">
        <div class="col">
          <h2>Modifier le participant</h2>
          <form method="post" action="/admin/member/update">
            <input type="hidden" name="member_id" value="{member_id}" />
            <label>Nom complet</label>
            <input name="full_name" value="{m['full_name']}" required />
            <div style="height:10px"></div>
            <label>Téléphone (9 chiffres)</label>
            <input name="phone" value="{m['phone']}" required />
            <div style="height:12px"></div>
            <button class="btn primary" type="submit">Enregistrer</button>
          </form>
        </div>

        <div class="col">
          <h2>Modifier PIN</h2>
          <form method="post" action="/admin/member/pin">
            <input type="hidden" name="member_id" value="{member_id}" />
            <label>Nouveau PIN (4 chiffres)</label>
            <input name="new_pin" placeholder="1234" required />
            <div style="height:12px"></div>
            <button class="btn primary" type="submit">Changer PIN</button>
          </form>
          <div style="height:14px"></div>
          {action_btn}
        </div>
      </div>
    </div>

    <div class="card">
      <h2>Enregistrer une cotisation</h2>
      <form method="post" action="/admin/add-payment">
        <input type="hidden" name="member_id" value="{member_id}" />

        <div class="row">
          <div class="col"><label>Date</label><input name="pay_date" value="{asof.isoformat()}" /></div>
          <div class="col"><label>Montant (F)</label><input name="amount" type="number" min="0" step="100" required /></div>
          <div class="col">
            <label>Mode</label>
            <select name="mode">
              <option value="wave">wave</option>
              <option value="om">om</option>
              <option value="cash">cash</option>
            </select>
          </div>
        </div>

        <div style="height:10px"></div>

        <div class="row">
          <div class="col"><label>Référence (optionnel)</label><input name="reference" placeholder="TXN..." /></div>
          <div class="col"><label>Note (optionnel)</label><input name="note" placeholder="cotisation..." /></div>
        </div>

        <div style="height:12px"></div>
        <button class="btn primary" type="submit">Ajouter</button>
        <a class="btn" href="/admin">Retour</a>
      </form>
    </div>

    <div class="card">
      <h2>Historique cotisations</h2>

      <div class="searchbar">
        <div style="flex:1; min-width:240px;">
          <label>Recherche (date / mode / montant)</label>
          <input placeholder="Tape ici..." data-search-table="#tblPays"/>
        </div>
      </div>

      <div class="table-wrap">
        <table id="tblPays" data-sortable>
          <thead><tr><th>Date</th><th>Montant</th><th>Mode</th><th>Référence</th><th>Note</th></tr></thead>
          <tbody>{''.join(pay_rows) if pay_rows else '<tr><td colspan="5">Aucune cotisation</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    """
    return render("Membre (Admin)", body, admin=True)


@app.post("/admin/member/update")
def admin_update_member(
    request: Request,
    member_id: int = Form(...),
    full_name: str = Form(...),
    phone: str = Form(...),
):
    gate = require_admin(request)
    if gate:
        return gate

    full_name = full_name.strip()
    phone = phone.strip()

    if not full_name:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Nom vide.</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    if not is_valid_sn_phone(phone):
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Téléphone invalide (9 chiffres).</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    with db() as con:
        m = con.execute("SELECT id FROM members WHERE id=?", (int(member_id),)).fetchone()
        if not m:
            return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Membre introuvable.</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

        other = con.execute("SELECT id FROM members WHERE phone=? AND id<>?", (phone, int(member_id))).fetchone()
        if other:
            return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Ce téléphone est déjà utilisé par un autre membre.</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

        con.execute("UPDATE members SET full_name=?, phone=? WHERE id=?", (full_name, phone, int(member_id)))

    return RedirectResponse(f"/admin/member/{member_id}", status_code=303)


@app.post("/admin/member/pin")
def admin_change_pin(
    request: Request,
    member_id: int = Form(...),
    new_pin: str = Form(...)
):
    gate = require_admin(request)
    if gate:
        return gate

    new_pin = new_pin.strip()
    if not (new_pin.isdigit() and len(new_pin) == 4):
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>PIN invalide (4 chiffres).</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    with db() as con:
        con.execute("UPDATE members SET pin_hash=? WHERE id=?", (pbkdf2_sha256.hash(new_pin), int(member_id)))

    return RedirectResponse(f"/admin/member/{member_id}", status_code=303)


@app.post("/admin/member/deactivate")
def admin_deactivate_member(request: Request, member_id: int = Form(...)):
    gate = require_admin(request)
    if gate:
        return gate
    with db() as con:
        con.execute("UPDATE members SET is_active=0 WHERE id=?", (int(member_id),))
    return RedirectResponse(f"/admin/member/{member_id}", status_code=303)


@app.post("/admin/member/activate")
def admin_activate_member(request: Request, member_id: int = Form(...)):
    gate = require_admin(request)
    if gate:
        return gate
    with db() as con:
        con.execute("UPDATE members SET is_active=1 WHERE id=?", (int(member_id),))
    return RedirectResponse(f"/admin/member/{member_id}", status_code=303)


@app.post("/admin/add-payment")
def admin_add_payment(
    request: Request,
    member_id: int = Form(...),
    pay_date: str = Form(...),
    amount: int = Form(...),
    mode: str = Form(...),
    reference: str = Form(""),
    note: str = Form("")
):
    gate = require_admin(request)
    if gate:
        return gate

    mode = mode.strip().lower()
    if mode not in {"wave", "om", "cash"}:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Mode invalide.</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    try:
        d = date.fromisoformat(pay_date.strip())
    except Exception:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Date invalide (AAAA-MM-JJ).</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    if amount <= 0:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Montant invalide.</p><a class="btn" href="/admin">Retour</a></div>', admin=True)

    with db() as con:
        con.execute(
            "INSERT INTO payments(member_id, pay_date, amount, mode, reference, note, created_at) VALUES(?,?,?,?,?,?,?)",
            (int(member_id), d.isoformat(), int(amount), mode, reference.strip(), note.strip(),
             datetime.now().isoformat(timespec="seconds")),
        )

    return RedirectResponse(f"/admin/member/{member_id}", status_code=303)


# ============================================================
# Retards
# ============================================================
@app.get("/admin/restes", response_class=HTMLResponse)
def admin_restes(request: Request):
    gate = require_admin(request)
    if gate:
        return gate

    asof = as_of_date()

    with db() as con:
        members = con.execute("SELECT * FROM members WHERE is_active=1").fetchall()

    data = []
    for m in members:
        m_start = date.fromisoformat(m["start_date"])
        st = member_status(int(m["id"]), m_start, asof)
        rest = int(st["rest"])
        if rest > 0:
            data.append((rest, m, st))

    data.sort(key=lambda x: x[0])

    rows = []
    for rest, m, st in data:
        rows.append(f"""
          <tr>
            <td>
              <a href="/admin/member/{m['id']}"><b>{m['full_name']}</b></a>
              <div class="muted">{m['phone']}</div>
            </td>
            <td class="mono">{fmt(int(st['paid']))} F</td>
            <td class="mono">{fmt(int(st['due']))} F</td>
            <td class="mono"><b>{fmt(int(rest))} F</b></td>
          </tr>
        """)

    body = f"""
    <div class="card">
      <h1>Participants en retard</h1>
      <div class="muted">Triés du plus petit au plus grand reste (début → {asof.isoformat()}).</div>

      <div class="searchbar">
        <div style="flex:1; min-width:240px;">
          <label>Recherche (nom / téléphone)</label>
          <input placeholder="Tape ici..." data-search-table="#tblRestes"/>
        </div>
      </div>

      <div class="table-wrap">
        <table id="tblRestes" data-sortable>
          <thead><tr><th>Participant</th><th>Payé</th><th>Dû</th><th>Reste</th></tr></thead>
          <tbody>{''.join(rows) if rows else '<tr><td colspan="4">Aucun retard 🎉</td></tr>'}</tbody>
        </table>
      </div>

      <div style="height:12px"></div>
      <a class="btn" href="/admin">Retour Dashboard</a>
    </div>
    """
    return render("Retards", body, admin=True)


# ============================================================
# Configuration + arrêt/reprise projet
# ============================================================
@app.get("/admin/config", response_class=HTMLResponse)
def admin_config(request: Request):
    gate = require_admin(request)
    if gate:
        return gate

    pstart = project_start_date().isoformat()
    pend = project_end_date().isoformat() if project_end_date() else ""
    da = daily_amount()
    mc = monthly_cap()
    octf = oct_2025_fixed()

    if not pend:
        stop_block = f"""
        <div class="card">
          <h2>Arrêter le projet</h2>
          <div class="muted">Fixe une date d'arrêt. Les calculs s'arrêtent à cette date.</div>
          <div style="height:10px"></div>
          <form method="post" action="/admin/project/stop" onsubmit="return confirm('Arrêter le projet ?');">
            <label>Date d'arrêt (AAAA-MM-JJ)</label>
            <input name="end_date" value="{as_of_date().isoformat()}" />
            <div style="height:12px"></div>
            <button class="btn danger" type="submit">Arrêter le projet</button>
          </form>
        </div>
        """
    else:
        stop_block = f"""
        <div class="card">
          <div class="warnbox">
            <h2>Projet arrêté</h2>
            <div class="muted">Date d'arrêt: <b>{pend}</b></div>
            <div style="height:12px"></div>
            <form method="post" action="/admin/project/resume" onsubmit="return confirm('Reprendre le projet ?');">
              <button class="btn" type="submit">Reprendre le projet</button>
            </form>
          </div>
        </div>
        """

    body = f"""
    <div class="card">
      <h1>Configuration du projet</h1>
      <form method="post" action="/admin/config/save">
        <div class="row">
          <div class="col">
            <label>Date début projet</label>
            <input name="project_start_date" value="{pstart}" />
          </div>
          <div class="col">
            <label>Montant journalier (F)</label>
            <input name="daily_amount_in" type="number" min="0" step="1" value="{da}" />
          </div>
          <div class="col">
            <label>Plafond mensuel (F)</label>
            <input name="monthly_cap_in" type="number" min="0" step="1" value="{mc}" />
          </div>
        </div>

        <div style="height:10px"></div>

        <div class="row">
          <div class="col">
            <label>Forfait Octobre 2025 (F)</label>
            <input name="oct_2025_fixed_in" type="number" min="0" step="1" value="{octf}" />
          </div>
          <div class="col">
            <label>Date d'arrêt actuelle (lecture)</label>
            <input value="{pend}" disabled />
          </div>
        </div>

        <div style="height:12px"></div>
        <button class="btn primary" type="submit">Enregistrer configuration</button>
        <a class="btn" href="/admin">Retour</a>
      </form>
    </div>

    {stop_block}
    """
    return render("Configuration", body, admin=True)


@app.post("/admin/config/save")
def admin_config_save(
    request: Request,
    project_start_date: str = Form(...),
    daily_amount_in: int = Form(...),
    monthly_cap_in: int = Form(...),
    oct_2025_fixed_in: int = Form(...),
):
    gate = require_admin(request)
    if gate:
        return gate

    try:
        _ = date.fromisoformat(project_start_date.strip())
    except Exception:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Date début invalide.</p><a class="btn" href="/admin/config">Retour</a></div>', admin=True)

    if daily_amount_in < 0 or monthly_cap_in < 0 or oct_2025_fixed_in < 0:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Montants invalides.</p><a class="btn" href="/admin/config">Retour</a></div>', admin=True)

    set_setting("project_start_date", project_start_date.strip())
    set_setting("daily_amount", str(int(daily_amount_in)))
    set_setting("monthly_cap", str(int(monthly_cap_in)))
    set_setting("oct_2025_fixed", str(int(oct_2025_fixed_in)))

    return RedirectResponse("/admin/config", status_code=303)


@app.post("/admin/project/stop")
def admin_project_stop(request: Request, end_date: str = Form(...)):
    gate = require_admin(request)
    if gate:
        return gate

    end_date = end_date.strip()
    try:
        d = date.fromisoformat(end_date)
    except Exception:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Date arrêt invalide (AAAA-MM-JJ).</p><a class="btn" href="/admin/config">Retour</a></div>', admin=True)

    set_setting("project_end_date", d.isoformat())
    return RedirectResponse("/admin/config", status_code=303)


@app.post("/admin/project/resume")
def admin_project_resume(request: Request):
    gate = require_admin(request)
    if gate:
        return gate
    set_setting("project_end_date", "")
    return RedirectResponse("/admin/config", status_code=303)


# ============================================================
# Résumé projet + Graphiques + Export Excel
# ============================================================
@app.get("/admin/resume", response_class=HTMLResponse)
def admin_resume(request: Request):
    gate = require_admin(request)
    if gate:
        return gate

    asof = as_of_date()
    collected = total_collected_as_of(asof)
    deposited = total_deposited_as_of(asof)
    caisse = max(0, collected - deposited)

    with db() as con:
        modes = con.execute(
            """
            SELECT
              COALESCE(SUM(amount),0) AS total,
              COALESCE(SUM(CASE WHEN mode='wave' THEN amount ELSE 0 END),0) AS wave_total,
              COALESCE(SUM(CASE WHEN mode='om' THEN amount ELSE 0 END),0) AS om_total,
              COALESCE(SUM(CASE WHEN mode='cash' THEN amount ELSE 0 END),0) AS cash_total
            FROM payments
            WHERE pay_date <= ?
            """,
            (asof.isoformat(),),
        ).fetchone()

    stats = monthly_collection_stats(asof)
    labels = list(stats.keys())
    totals = [stats[k]["total"] for k in labels]
    wave = [stats[k]["wave"] for k in labels]
    om = [stats[k]["om"] for k in labels]
    cash = [stats[k]["cash"] for k in labels]

    js_data = json.dumps({"labels": labels, "totals": totals, "wave": wave, "om": om, "cash": cash})

    body = f"""
    <div class="card">
      <h1>Résumé du projet</h1>
      <div class="muted">Calcul arrêté à: <b>{asof.isoformat()}</b></div>

      <div style="height:14px"></div>
      <div class="kpis">
        <div class="kpi"><div class="t">Total encaissé</div><div class="v">{fmt(collected)} F</div></div>
        <div class="kpi"><div class="t">Total versé banque</div><div class="v">{fmt(deposited)} F</div></div>
        <div class="kpi"><div class="t">Caisse (dispo à verser)</div><div class="v">{fmt(caisse)} F</div></div>
        <div class="kpi"><div class="t">Export Excel</div><div class="v"><a href="/admin/resume/export.xlsx">Télécharger</a></div></div>
      </div>

      <div style="height:14px"></div>
      <div class="kpis">
        <div class="kpi"><div class="t">Wave (global)</div><div class="v">{fmt(int(modes["wave_total"]))} F</div></div>
        <div class="kpi"><div class="t">OM (global)</div><div class="v">{fmt(int(modes["om_total"]))} F</div></div>
        <div class="kpi"><div class="t">Cash (global)</div><div class="v">{fmt(int(modes["cash_total"]))} F</div></div>
        <div class="kpi"><div class="t">Comptabilité</div><div class="v"><a href="/admin/comptabilite">Ouvrir</a></div></div>
      </div>
    </div>

    <div class="card">
      <h2>Graphiques</h2>
      <div class="muted">Encaissé par mois + Wave/OM/Cash.</div>
      <div style="height:10px"></div>

      <div class="row">
        <div class="col">
          <h2>Encaissé par mois (Total)</h2>
          <canvas id="c1" width="900" height="320" style="width:100%; border:1px solid rgba(255,255,255,.08); border-radius:16px;"></canvas>
        </div>
      </div>

      <div style="height:12px"></div>
      <div class="row">
        <div class="col">
          <h2>Wave / OM / Cash par mois</h2>
          <canvas id="c2" width="900" height="360" style="width:100%; border:1px solid rgba(255,255,255,.08); border-radius:16px;"></canvas>
        </div>
      </div>
    </div>

    <script>
      const DATA = {js_data};

      function drawBars(canvasId, labels, series){{
        const c = document.getElementById(canvasId);
        const ctx = c.getContext("2d");
        const W = c.width, H = c.height;
        ctx.clearRect(0,0,W,H);

        const pad = 40;
        const innerW = W - pad*2;
        const innerH = H - pad*2;

        const all = [];
        series.forEach(s=>s.data.forEach(v=>all.push(v)));
        const maxV = Math.max(1, ...all);

        // axes
        ctx.lineWidth = 1;
        ctx.strokeStyle = "rgba(255,255,255,.18)";
        if(document.body.getAttribute("data-theme")==="light") ctx.strokeStyle = "rgba(10,16,32,.18)";
        ctx.beginPath();
        ctx.moveTo(pad, pad);
        ctx.lineTo(pad, H-pad);
        ctx.lineTo(W-pad, H-pad);
        ctx.stroke();

        const n = labels.length;
        const groups = series.length;
        const gap = 10;
        const groupW = innerW / Math.max(1,n);
        const barW = Math.max(4, (groupW - gap) / Math.max(1,groups));

        // grid + y labels
        ctx.font = "12px system-ui";
        for(let i=0;i<=4;i++) {{
          const y = pad + (innerH*(i/4));
          ctx.strokeStyle = "rgba(255,255,255,.08)";
          if(document.body.getAttribute("data-theme")==="light") ctx.strokeStyle = "rgba(10,16,32,.08)";
          ctx.beginPath();
          ctx.moveTo(pad, y);
          ctx.lineTo(W-pad, y);
          ctx.stroke();

          const val = Math.round(maxV*(1 - i/4));
          ctx.fillStyle = (document.body.getAttribute("data-theme")==="light") ? "rgba(10,16,32,.80)" : "rgba(255,255,255,.80)";
          ctx.fillText(val.toString(), 6, y+4);
        }}

        // bars
        series.forEach((s, si)=>{{
          for(let i=0;i<n;i++) {{
            const v = s.data[i] || 0;
            const h = (v/maxV) * innerH;
            const x = pad + i*groupW + si*barW + (gap/2);
            const y = (H-pad) - h;

            const isLight = (document.body.getAttribute("data-theme") === "light");
            ctx.fillStyle = isLight ? (s.light || "rgba(10,16,32,.75)") : (s.dark || "rgba(255,255,255,.75)");
            ctx.globalAlpha = 0.9;
            ctx.fillRect(x, y, barW-2, h);
          }}
        }});

        // x labels
        ctx.globalAlpha = 0.95;
        ctx.fillStyle = (document.body.getAttribute("data-theme")==="light") ? "rgba(10,16,32,.85)" : "rgba(255,255,255,.85)";
        ctx.font = "11px system-ui";
        const step = n > 16 ? 2 : 1;
        for(let i=0;i<n;i+=step) {{
          const x = pad + i*groupW + 2;
          ctx.fillText(labels[i], x, H-12);
        }}
      }}

      function renderCharts(){{
        drawBars("c1", DATA.labels, [
          {{ name:"Total", data: DATA.totals, dark:"rgba(106,168,255,.85)", light:"rgba(55,120,255,.85)" }}
        ]);

        drawBars("c2", DATA.labels, [
          {{ name:"Wave", data: DATA.wave, dark:"rgba(45,212,191,.85)", light:"rgba(12,148,130,.85)" }},
          {{ name:"OM", data: DATA.om, dark:"rgba(155,123,255,.85)", light:"rgba(120,80,230,.85)" }},
          {{ name:"Cash", data: DATA.cash, dark:"rgba(251,191,36,.85)", light:"rgba(210,140,10,.85)" }},
        ]);
      }}

      renderCharts();

      // re-render when theme changes
      const _oldToggle = window.toggleTheme;
      window.toggleTheme = function(){{
        _oldToggle();
        setTimeout(renderCharts, 100);
      }};
    </script>
    """
    return render("Résumé", body, admin=True)


@app.get("/admin/resume/export.xlsx")
def admin_resume_export_xlsx(request: Request):
    gate = require_admin(request)
    if gate:
        return PlainTextResponse("Accès refusé.", status_code=403)

    try:
        from openpyxl import Workbook
    except Exception:
        return PlainTextResponse("openpyxl manquant. Installe: pip install openpyxl", status_code=500)

    asof = as_of_date()
    collected = total_collected_as_of(asof)
    deposited = total_deposited_as_of(asof)
    caisse = max(0, collected - deposited)

    stats = monthly_collection_stats(asof)
    labels = list(stats.keys())

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resume"
    ws1.append(["AsOf", asof.isoformat()])
    ws1.append(["Total encaisse", collected])
    ws1.append(["Total verse banque", deposited])
    ws1.append(["Caisse (dispo a verser)", caisse])

    ws2 = wb.create_sheet("Encaisse_mensuel")
    ws2.append(["Mois", "Total", "Wave", "OM", "Cash"])
    for ym in labels:
        ws2.append([ym, stats[ym]["total"], stats[ym]["wave"], stats[ym]["om"], stats[ym]["cash"]])

    filename = f"tontine_resume_{asof.isoformat()}.xlsx"
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return Response(
        content=bio.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
# Comptabilité + Banque + Supprimer paiement + Exports
# ============================================================
@app.get("/admin/comptabilite", response_class=HTMLResponse)
def admin_comptabilite(request: Request, mois: int = 0, annee: int = 0):
    gate = require_admin(request)
    if gate:
        return gate

    t = as_of_date()
    if mois == 0:
        mois = t.month
    if annee == 0:
        annee = t.year

    ym = f"{annee:04d}-{mois:02d}"

    with db() as con:
        members = con.execute(
            "SELECT id, full_name, phone FROM members WHERE is_active=1 ORDER BY full_name"
        ).fetchall()

        journal = con.execute(
            """
            SELECT p.id, p.pay_date, p.amount, p.mode, p.reference, p.note, m.full_name, m.phone
            FROM payments p
            JOIN members m ON m.id = p.member_id
            WHERE substr(p.pay_date, 1, 7) = ?
            ORDER BY p.pay_date DESC, p.id DESC
            """,
            (ym,),
        ).fetchall()

        totals = con.execute(
            """
            SELECT
              COALESCE(SUM(amount),0) AS total,
              COALESCE(SUM(CASE WHEN mode='wave' THEN amount ELSE 0 END),0) AS wave_total,
              COALESCE(SUM(CASE WHEN mode='om' THEN amount ELSE 0 END),0) AS om_total,
              COALESCE(SUM(CASE WHEN mode='cash' THEN amount ELSE 0 END),0) AS cash_total
            FROM payments
            WHERE substr(pay_date, 1, 7) = ?
            """,
            (ym,),
        ).fetchone()

        bank_rows = con.execute(
            """
            SELECT id, dep_date, amount, reference, note
            FROM bank_deposits
            WHERE substr(dep_date, 1, 7) = ?
            ORDER BY dep_date DESC, id DESC
            """,
            (ym,),
        ).fetchall()

        bank_totals = con.execute(
            "SELECT COALESCE(SUM(amount),0) AS total FROM bank_deposits WHERE substr(dep_date, 1, 7) = ?",
            (ym,),
        ).fetchone()

    options = "\n".join(
        [f'<option value="{m["id"]}">{m["full_name"]} ({m["phone"]})</option>' for m in members]
    )

    asof = as_of_date()
    collected_all = total_collected_as_of(asof)
    deposited_all = total_deposited_as_of(asof)
    available_bank = max(0, collected_all - deposited_all)

    rows = []
    for j in journal:
        rows.append(f"""
          <tr>
            <td class="mono">{j['pay_date']}</td>
            <td><b>{j['full_name']}</b><div class="muted">{j['phone']}</div></td>
            <td class="mono"><b>{fmt(int(j['amount']))} F</b></td>
            <td>{j['mode']}</td>
            <td class="muted">{j['reference'] or ""}</td>
            <td class="muted">{j['note'] or ""}</td>
            <td>
              <form method="post" action="/admin/payment/delete" onsubmit="return confirm('Supprimer ce paiement ?');">
                <input type="hidden" name="payment_id" value="{j['id']}" />
                <input type="hidden" name="mois" value="{mois}" />
                <input type="hidden" name="annee" value="{annee}" />
                <button class="btn danger" type="submit">Supprimer</button>
              </form>
            </td>
          </tr>
        """)

    bank_table_rows = []
    for b in bank_rows:
        bank_table_rows.append(f"""
          <tr>
            <td class="mono">{b['dep_date']}</td>
            <td class="mono"><b>{fmt(int(b['amount']))} F</b></td>
            <td class="muted">{b['reference'] or ""}</td>
            <td class="muted">{b['note'] or ""}</td>
            <td>
              <form method="post" action="/admin/bank/delete" onsubmit="return confirm('Supprimer ce versement banque ?');">
                <input type="hidden" name="deposit_id" value="{b['id']}" />
                <input type="hidden" name="mois" value="{mois}" />
                <input type="hidden" name="annee" value="{annee}" />
                <button class="btn danger" type="submit">Supprimer</button>
              </form>
            </td>
          </tr>
        """)

    body = f"""
    <div class="card">
      <h1>Comptabilité ({ym})</h1>

      <form method="get" action="/admin/comptabilite" class="row">
        <div class="col">
          <label>Mois</label>
          <input type="number" name="mois" min="1" max="12" value="{mois}" />
        </div>
        <div class="col">
          <label>Année</label>
          <input type="number" name="annee" min="2020" max="2100" value="{annee}" />
        </div>
        <div class="col" style="align-self:end;">
          <button class="btn" type="submit">Filtrer</button>
          <a class="btn" href="/admin/comptabilite/export.csv?mois={mois}&annee={annee}">Exporter CSV</a>
          <a class="btn" href="/admin/comptabilite/export.xlsx?mois={mois}&annee={annee}">Exporter Excel</a>
        </div>
      </form>

      <div style="height:12px"></div>
      <div class="kpis">
        <div class="kpi"><div class="t">Total encaissé (mois)</div><div class="v">{fmt(int(totals['total']))} F</div></div>
        <div class="kpi"><div class="t">Wave (mois)</div><div class="v">{fmt(int(totals['wave_total']))} F</div></div>
        <div class="kpi"><div class="t">OM (mois)</div><div class="v">{fmt(int(totals['om_total']))} F</div></div>
        <div class="kpi"><div class="t">Cash (mois)</div><div class="v">{fmt(int(totals['cash_total']))} F</div></div>
      </div>

      <div style="height:12px"></div>
      <div class="kpis">
        <div class="kpi"><div class="t">Total encaissé (début → {asof.isoformat()})</div><div class="v">{fmt(collected_all)} F</div></div>
        <div class="kpi"><div class="t">Total versé banque</div><div class="v">{fmt(deposited_all)} F</div></div>
        <div class="kpi"><div class="t">Solde dispo à verser</div><div class="v">{fmt(available_bank)} F</div></div>
        <div class="kpi"><div class="t">Total versé banque ({ym})</div><div class="v">{fmt(int(bank_totals['total']))} F</div></div>
      </div>
    </div>

    <div class="card">
      <h2>Ajouter une cotisation</h2>
      <form method="post" action="/admin/comptabilite/add">
        <div class="row">
          <div class="col">
            <label>Date</label>
            <input name="pay_date" value="{t.isoformat()}" />
          </div>
          <div class="col">
            <label>Membre</label>
            <select name="member_id" required>
              {options}
            </select>
          </div>
        </div>

        <div style="height:10px"></div>

        <div class="row">
          <div class="col">
            <label>Montant (F)</label>
            <input name="amount" type="number" min="0" step="100" required />
          </div>
          <div class="col">
            <label>Mode</label>
            <select name="mode">
              <option value="wave">wave</option>
              <option value="om">om</option>
              <option value="cash">cash</option>
            </select>
          </div>
        </div>

        <div style="height:10px"></div>

        <div class="row">
          <div class="col">
            <label>Référence (optionnel)</label>
            <input name="reference" placeholder="TXN..." />
          </div>
          <div class="col">
            <label>Note (optionnel)</label>
            <input name="note" placeholder="cotisation..." />
          </div>
        </div>

        <input type="hidden" name="mois" value="{mois}" />
        <input type="hidden" name="annee" value="{annee}" />

        <div style="height:12px"></div>
        <button class="btn primary" type="submit">Enregistrer</button>
      </form>
    </div>

    <div class="card">
      <h2>Journal paiements ({ym})</h2>

      <div class="searchbar">
        <div style="flex:1; min-width:240px;">
          <label>Recherche (nom / téléphone / mode / montant)</label>
          <input placeholder="Tape ici..." data-search-table="#tblJournal"/>
        </div>
      </div>

      <div class="table-wrap">
        <table id="tblJournal" data-sortable>
          <thead><tr><th>Date</th><th>Membre</th><th>Montant</th><th>Mode</th><th>Réf</th><th>Note</th><th>Actions</th></tr></thead>
          <tbody>{''.join(rows) if rows else '<tr><td colspan="7">Aucune écriture</td></tr>'}</tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <h2>Versements Banque</h2>
      <div class="muted">Disponible à verser maintenant: <b>{fmt(available_bank)} F</b></div>

      <div style="height:10px"></div>
      <form method="post" action="/admin/bank/add">
        <div class="row">
          <div class="col"><label>Date</label><input name="dep_date" value="{t.isoformat()}" /></div>
          <div class="col"><label>Montant (F)</label><input name="amount" type="number" min="0" step="100" required /></div>
        </div>

        <div style="height:10px"></div>
        <div class="row">
          <div class="col"><label>Référence (optionnel)</label><input name="reference" placeholder="BORDEREAU..." /></div>
          <div class="col"><label>Note (optionnel)</label><input name="note" placeholder="Versement banque..." /></div>
        </div>

        <input type="hidden" name="mois" value="{mois}" />
        <input type="hidden" name="annee" value="{annee}" />

        <div style="height:12px"></div>
        <button class="btn primary" type="submit">Enregistrer versement</button>
      </form>

      <div style="height:14px"></div>

      <div class="searchbar">
        <div style="flex:1; min-width:240px;">
          <label>Recherche (date / montant / note)</label>
          <input placeholder="Tape ici..." data-search-table="#tblBank"/>
        </div>
      </div>

      <div class="table-wrap">
        <table id="tblBank" data-sortable>
          <thead><tr><th>Date</th><th>Montant</th><th>Réf</th><th>Note</th><th>Actions</th></tr></thead>
          <tbody>{''.join(bank_table_rows) if bank_table_rows else '<tr><td colspan="5">Aucun versement banque</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    """
    return render("Comptabilité", body, admin=True)


@app.post("/admin/comptabilite/add")
def admin_compta_add(
    request: Request,
    pay_date: str = Form(...),
    member_id: int = Form(...),
    amount: int = Form(...),
    mode: str = Form(...),
    reference: str = Form(""),
    note: str = Form(""),
    mois: int = Form(0),
    annee: int = Form(0),
):
    gate = require_admin(request)
    if gate:
        return gate

    mode = mode.strip().lower()
    if mode not in {"wave", "om", "cash"}:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Mode invalide.</p><a class="btn" href="/admin/comptabilite">Retour</a></div>', admin=True)

    try:
        d = date.fromisoformat(pay_date.strip())
    except Exception:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Date invalide (AAAA-MM-JJ).</p><a class="btn" href="/admin/comptabilite">Retour</a></div>', admin=True)

    if amount <= 0:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Montant invalide.</p><a class="btn" href="/admin/comptabilite">Retour</a></div>', admin=True)

    with db() as con:
        con.execute(
            "INSERT INTO payments(member_id, pay_date, amount, mode, reference, note, created_at) VALUES(?,?,?,?,?,?,?)",
            (int(member_id), d.isoformat(), int(amount), mode, reference.strip(), note.strip(),
             datetime.now().isoformat(timespec="seconds")),
        )

    if mois == 0:
        mois = d.month
    if annee == 0:
        annee = d.year

    return RedirectResponse(f"/admin/comptabilite?mois={mois}&annee={annee}", status_code=303)


@app.post("/admin/payment/delete")
def admin_delete_payment(
    request: Request,
    payment_id: int = Form(...),
    mois: int = Form(0),
    annee: int = Form(0),
):
    gate = require_admin(request)
    if gate:
        return gate

    with db() as con:
        con.execute("DELETE FROM payments WHERE id=?", (int(payment_id),))

    if mois and annee:
        return RedirectResponse(f"/admin/comptabilite?mois={mois}&annee={annee}", status_code=303)
    return RedirectResponse("/admin/comptabilite", status_code=303)


@app.post("/admin/bank/add")
def admin_add_bank_deposit(
    request: Request,
    dep_date: str = Form(...),
    amount: int = Form(...),
    reference: str = Form(""),
    note: str = Form(""),
    mois: int = Form(0),
    annee: int = Form(0),
):
    gate = require_admin(request)
    if gate:
        return gate

    try:
        d = date.fromisoformat(dep_date.strip())
    except Exception:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Date banque invalide (AAAA-MM-JJ).</p><a class="btn" href="/admin/comptabilite">Retour</a></div>', admin=True)

    if amount <= 0:
        return render("Erreur", '<div class="card"><h1>Erreur</h1><p>Montant banque invalide.</p><a class="btn" href="/admin/comptabilite">Retour</a></div>', admin=True)

    asof = as_of_date()
    collected = total_collected_as_of(asof)
    deposited = total_deposited_as_of(asof)
    available = max(0, collected - deposited)

    if amount > available:
        return render(
            "Erreur",
            f'<div class="card"><h1>Erreur</h1>'
            f'<p>Montant &gt; solde disponible.</p>'
            f'<p>Disponible: <b>{fmt(available)} F</b></p>'
            f'<a class="btn" href="/admin/comptabilite">Retour</a></div>',
            admin=True,
        )

    with db() as con:
        con.execute(
            "INSERT INTO bank_deposits(dep_date, amount, reference, note, created_at) VALUES(?,?,?,?,?)",
            (d.isoformat(), int(amount), reference.strip(), note.strip(), datetime.now().isoformat(timespec="seconds")),
        )

    if mois == 0:
        mois = d.month
    if annee == 0:
        annee = d.year
    return RedirectResponse(f"/admin/comptabilite?mois={mois}&annee={annee}", status_code=303)


@app.post("/admin/bank/delete")
def admin_delete_bank_deposit(
    request: Request,
    deposit_id: int = Form(...),
    mois: int = Form(0),
    annee: int = Form(0),
):
    gate = require_admin(request)
    if gate:
        return gate

    with db() as con:
        con.execute("DELETE FROM bank_deposits WHERE id=?", (int(deposit_id),))

    if mois and annee:
        return RedirectResponse(f"/admin/comptabilite?mois={mois}&annee={annee}", status_code=303)
    return RedirectResponse("/admin/comptabilite", status_code=303)


@app.get("/admin/comptabilite/export.csv", response_class=PlainTextResponse)
def admin_comptabilite_export_csv(request: Request, mois: int, annee: int):
    gate = require_admin(request)
    if gate:
        return PlainTextResponse("Accès refusé.", status_code=403)

    ym = f"{annee:04d}-{mois:02d}"

    with db() as con:
        journal = con.execute(
            """
            SELECT p.pay_date, m.full_name, m.phone, p.amount, p.mode, p.reference, p.note
            FROM payments p
            JOIN members m ON m.id = p.member_id
            WHERE substr(p.pay_date, 1, 7) = ?
            ORDER BY p.pay_date ASC, p.id ASC
            """,
            (ym,),
        ).fetchall()

    lines = ["date;nom;telephone;montant;mode;reference;note"]
    for j in journal:
        lines.append(
            f'{j["pay_date"]};{(j["full_name"] or "").replace(";", ",")};{j["phone"]};{int(j["amount"])};{j["mode"]};{(j["reference"] or "").replace(";", ",")};{(j["note"] or "").replace(";", ",")}'
        )
    return PlainTextResponse("\n".join(lines), media_type="text/csv; charset=utf-8")


@app.get("/admin/comptabilite/export.xlsx")
def admin_comptabilite_export_xlsx(request: Request, mois: int, annee: int):
    gate = require_admin(request)
    if gate:
        return PlainTextResponse("Accès refusé.", status_code=403)

    try:
        from openpyxl import Workbook
    except Exception:
        return PlainTextResponse("openpyxl manquant. Installe: pip install openpyxl", status_code=500)

    ym = f"{annee:04d}-{mois:02d}"
    with db() as con:
        journal = con.execute(
            """
            SELECT p.pay_date, m.full_name, m.phone, p.amount, p.mode, p.reference, p.note
            FROM payments p
            JOIN members m ON m.id = p.member_id
            WHERE substr(p.pay_date, 1, 7) = ?
            ORDER BY p.pay_date ASC, p.id ASC
            """,
            (ym,),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Paiements_{ym}"
    ws.append(["Date", "Nom", "Telephone", "Montant", "Mode", "Reference", "Note"])
    for j in journal:
        ws.append([j["pay_date"], j["full_name"], j["phone"], int(j["amount"]), j["mode"], j["reference"], j["note"]])

    filename = f"tontine_paiements_{ym}.xlsx"
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return Response(
        content=bio.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
# Member space
# ============================================================
@app.get("/me", response_class=HTMLResponse)
def me(request: Request):
    gate, member_id = require_member(request)
    if gate:
        return gate

    asof = as_of_date()

    with db() as con:
        m = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()

    if not m or int(m["is_active"]) != 1:
        resp = RedirectResponse("/", status_code=303)
        clear_cookie(resp, "member_id")
        return resp

    with db() as con:
        pays = con.execute(
            "SELECT * FROM payments WHERE member_id=? ORDER BY pay_date DESC, id DESC LIMIT 400",
            (member_id,),
        ).fetchall()

    m_start = date.fromisoformat(m["start_date"])
    st = member_status(member_id, m_start, asof)
    badge_cls = "ok" if st["status"] == "À jour" else "bad"

    rows = []
    for p in pays:
        rows.append(f"""
          <tr>
            <td class="mono">{p['pay_date']}</td>
            <td class="mono"><b>{fmt(int(p['amount']))} F</b></td>
            <td>{p['mode']}</td>
            <td class="muted">{p['reference'] or ""}</td>
          </tr>
        """)

    body = f"""
    <div class="card">
      <h1>Mon profil</h1>
      <div><b>{m['full_name']}</b> <span class="muted">({m['phone']})</span></div>
      <div class="muted">Début: <b>{m['start_date']}</b></div>
      <div style="height:10px"></div>
      <span class="badge {badge_cls}"><span class="b"></span>{st['status']}</span>

      <div style="height:14px"></div>
      <div class="kpis">
        <div class="kpi"><div class="t">Payé</div><div class="v">{fmt(int(st['paid']))} F</div></div>
        <div class="kpi"><div class="t">Dû (début → {asof.isoformat()})</div><div class="v">{fmt(int(st['due']))} F</div></div>
        <div class="kpi"><div class="t">Reste</div><div class="v">{fmt(int(st['rest']))} F</div></div>
        <div class="kpi"><div class="t">Avance</div><div class="v">{fmt(int(st['avance']))} F</div></div>
      </div>
    </div>

    <div class="card">
      <h2>Mes cotisations</h2>

      <div class="searchbar">
        <div style="flex:1; min-width:240px;">
          <label>Recherche (date / mode / montant)</label>
          <input placeholder="Tape ici..." data-search-table="#tblMe"/>
        </div>
      </div>

      <div class="table-wrap">
        <table id="tblMe" data-sortable>
          <thead><tr><th>Date</th><th>Montant</th><th>Mode</th><th>Référence</th></tr></thead>
          <tbody>{''.join(rows) if rows else '<tr><td colspan="4">Aucune cotisation</td></tr>'}</tbody>
        </table>
      </div>
    </div>
    """
    return render("Mon profil", body, member=True)
